import pandas as pd
from openpyxl import load_workbook

def insert_row(new_row: list, new_index: int, df: pd.DataFrame, total_overtime=[0,0], total_hours=[0,0]):
    if len(new_row) == 8:
        if (new_row[0] == "Celkem"):
            if total_overtime[1] < 10:
                new_row[5] = f"{total_overtime[0]}:0{total_overtime[1]}"
            if total_hours[1] < 10:
                new_row[3] = f"{total_hours[0]}:0{total_hours[1]}"
        else:
            new_row[1] = new_row[1][0:-2]
        df.loc[new_index] = new_row
    else:
        print("bad length of list")

def calculate_hours(hours: int, minutes: int):
    plus_hours = minutes / 60
    hours += int(plus_hours)
    minutes = round(float(plus_hours % 1 * 60))
    return [hours, minutes]

def convert(filename, output_filename):
    wb = load_workbook(filename = f"static/input/{filename}", data_only=True)

    writer = pd.ExcelWriter(f"static/output/{output_filename}", engine="xlsxwriter")
    workbook = writer.book
    format_yellow = workbook.add_format({'bg_color': 'yellow'})

    log = []

    for sheet in wb.sheetnames:
        if sheet != "Worksheet":
            #variables
            log_attendance = ""
            log_hours = ""
            indexes = []
            yellow_indexes = []
            total_hours = [0, 0]
            total_overtime = [0, 0]

            #wb deleting cols that are useless
            wb[sheet].delete_cols(4)
            wb[sheet].delete_cols(8)

            df = pd.DataFrame(wb[sheet].values)
            df = df.replace("Actual", "Práce").replace("Nelze zaplatit", "Přesčas")
            
            #collecting indexes that needs to be deleted and searching for halfday holiday
            for i in range(len(df.index)):
                if df.iloc[i][0] == "Čas začátku / konce" or df.iloc[i][0] == "Illness / Nemocenská" or df.iloc[i][0] == "Dovolená" or df.iloc[i][2] == "Celkem" or df.iloc[i][0] == "Stav pružné pracovní doby" or df.iloc[i][0] == "Obnovení pružné pracovní doby" or df.iloc[i][0] == "Přihlášen" or df.iloc[i][0] == None or df.iloc[i][0] == "Celkem" or df.iloc[i][0] == "Unpaid Leave / Neplacené volno" or df.iloc[i][0] == "Blood Donation / Darování krve":
                    indexes.append(i)

                if df.iloc[i][0] != None:
                    if "Den" in df.iloc[i][0]:
                        if i != 5:
                            indexes.append(i)

                if df.iloc[i][7] == "Dovolená" and df.iloc[i][4] == "4:00":
                    df.at[i, 7] = "Půl dne dovolená"
                    
            df = df.drop(indexes)

            #calculating total hours worked and overtime hours
            for i in range(len(df.index)):
                if df.iloc[i][3] != "None" and type(df.iloc[i][3]) == str and df.iloc[i][3] != "Práce" and df.iloc[i][3] != "---":
                    total_hours[0] += int(df.iloc[i][3][0:2].replace(":", ""))
                    total_hours[1] += int(df.iloc[i][3][2:5].replace(":", ""))
                    total_overtime[0] += int(df.iloc[i][5][0:2].replace(":", ""))
                    total_overtime[1] += int(df.iloc[i][5][2:5].replace(":", ""))

                    if int(df.iloc[i][3][0:2].replace(":", "")) >= 14:
                        log_hours += f"{df.iloc[i][0][0:3]},"

                    #if str(df.iloc[i][6])[8:10] == "16":
                    #    log_attendance += f"{df.iloc[i][0][0:3]}, "

                    if df.iloc[i][3] == "0:00" and df.iloc[i][4] == "8:00":
                        log_attendance += f"{df.iloc[i][0][0:3]}, "

            total_hours = calculate_hours(total_hours[0], total_hours[1])
            total_overtime = calculate_hours(total_overtime[0], total_overtime[1])

            #subsetting dataframe with dates and sorting them
            df_subset = df[4:len(df.index)].copy()
            df.drop(df.index[4:len(df.index)], axis=0, inplace=True)
            df_subset.sort_values(df_subset.columns[0], ascending=True, inplace=True)
            df = pd.concat([df, df_subset], ignore_index=True)

            #inserting new row with total time
            insert_row(["Celkem", "", "", f"{total_hours[0]}:{total_hours[1]}", "", f"{total_overtime[0]}:{total_overtime[1]}", "", ""], -1, df=df, total_overtime=total_overtime, total_hours=total_hours)
            
            #indexes for coloring
            for i in range(len(df.index)):
                if df.iloc[i][4] == "0:00" and df.iloc[i][7] == None:
                    yellow_indexes.append(i) 
            
            df.drop([4, 6], axis=1, inplace=True)

            df.to_excel(writer, sheet_name=sheet, header=False, index=False)

            #setting color for holiday and weekends and resizing columns
            worksheet = writer.sheets[sheet]
            worksheet.set_column(0, 3, 11)
            worksheet.set_column(3, 5, 6)
            worksheet.set_column(5, 6, 30)
            for i in yellow_indexes:
                for j in range(6):
                    worksheet.write(i, j, df.iloc[i, j], format_yellow)

            #logging
            log.append([sheet, log_attendance, log_hours])

        else:
            df_log = pd.DataFrame(log, columns=["Name", "Missed attendance", "Too many hours"])
            df_log.to_excel(writer, sheet_name="Log")
            worksheet = writer.sheets["Log"]
            worksheet.set_column(1, 1, 15)
            worksheet.set_column(2, 2, 70)
            worksheet.set_column(3, 3, 40)
    writer.close()

